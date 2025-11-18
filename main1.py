
import os
import sys
import logging
import tempfile
import shutil
import base64
from functools import wraps
from datetime import datetime
import re
import locale
from io import BytesIO

from dotenv import load_dotenv
from telegram import Update
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    filters
)
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIGURATION (từ config.py)
# ============================================================================

# Lấy đường dẫn thư mục hiện tại của script
current_dir = os.path.dirname(os.path.abspath(__file__))

# Tải biến môi trường từ tệp .env
dotenv_path = os.path.join(current_dir, '.env')
load_dotenv(dotenv_path)

# Thông tin của bot
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
EXCEL_TEMPLATE_BASE64 = os.getenv("EXCEL_TEMPLATE_BASE64")
BANGLUONG = os.getenv("BANGLUONG")  # Dữ liệu bảng lương dạng base64

# Danh sách user ID được phép sử dụng bot
ALLOWED_USERS_STR = os.getenv("ALLOWED_USERS", "")
if not ALLOWED_USERS_STR:
    print("CẢNH BÁO: ALLOWED_USERS không được cấu hình! Bot sẽ cho phép tất cả người dùng truy cập.")
    ALLOWED_USERS = []
else:
    try:
        ALLOWED_USERS = [int(id.strip()) for id in ALLOWED_USERS_STR.split(",") if id.strip()]
        if not ALLOWED_USERS:
            print("CẢNH BÁO: ALLOWED_USERS không chứa ID hợp lệ nào! Bot sẽ cho phép tất cả người dùng truy cập.")
    except ValueError as e:
        print(f"LỖI: Định dạng ALLOWED_USERS không hợp lệ! Bot sẽ cho phép tất cả người dùng truy cập. Chi tiết lỗi: {e}")
        ALLOWED_USERS = []

# Cấu hình network và file
NETWORK_TIMEOUT = int(os.getenv("NETWORK_TIMEOUT", "60"))  # seconds
MAX_RETRIES = int(os.getenv("MAX_RETRIES", "3"))
RETRY_DELAY = int(os.getenv("RETRY_DELAY", "2"))  # seconds
MAX_FILE_SIZE_MB = int(os.getenv("MAX_FILE_SIZE_MB", "50"))  # MB

# Kiểm tra các biến môi trường cần thiết
if not TELEGRAM_TOKEN:
    print("❌ LỖI: TELEGRAM_TOKEN không được tìm thấy! Vui lòng kiểm tra tệp .env.")

if not EXCEL_TEMPLATE_BASE64:
    print("❌ LỖI: EXCEL_TEMPLATE_BASE64 không được tìm thấy! Vui lòng kiểm tra tệp .env.")

# Validate network configurations
if NETWORK_TIMEOUT < 10:
    print("CẢNH BÁO: NETWORK_TIMEOUT quá thấp, đặt về 60 giây.")
    NETWORK_TIMEOUT = 60

if MAX_RETRIES < 1:
    print("CẢNH BÁO: MAX_RETRIES quá thấp, đặt về 3.")
    MAX_RETRIES = 3

if MAX_FILE_SIZE_MB > 100:
    print("CẢNH BÁO: MAX_FILE_SIZE_MB quá cao, đặt về 50MB.")
    MAX_FILE_SIZE_MB = 50

# ============================================================================
# EXCEL UTILITIES (từ excel_utils.py)
# ============================================================================

# Thiết lập logging cho excel utilities
logger = logging.getLogger(__name__)

def apply_cell_style(cell, font=None, alignment=None, border=None, number_format=None, fill=None):
    """Áp dụng style cho một ô."""
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    if fill:
        cell.fill = fill

def auto_adjust_column_width(worksheet):
    """Tự động điều chỉnh độ rộng của các cột."""
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        # Xét độ dài của tiêu đề
        header_cell = worksheet[f"{column_letter}1"]
        if header_cell.value:
            max_length = max(max_length, len(str(header_cell.value)))
        
        # Xét độ dài của các ô trong cột
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        # Điều chỉnh độ rộng cột, cộng thêm 2 để đệm
        worksheet.column_dimensions[column_letter].width = max_length + 2

def process_excel_file(input_file_path, output_file_path):
    """Xử lý file Excel đơn và tạo ra báo cáo định dạng."""
    try:
        # Tạo styles cho định dạng
        font_style = Font(name="Calibri", size=12)
        bold_font = Font(name="Calibri", bold=True, size=12)
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Xử lý file Excel đầu vào
        workbook = load_workbook(filename=input_file_path, data_only=True)
        sheet = workbook.active

        # Tìm vị trí các cột (dựa vào header)
        header = [cell.value for cell in sheet[1]]
        
        # Danh sách lưu các cột thiếu
        missing_columns = []
        
        try:
            customer_col_index = header.index("Khách hàng")
        except ValueError:
            missing_columns.append("Khách hàng")
            
        try:
            total_col_index = header.index("Khách cần trả")
        except ValueError:
            missing_columns.append("Khách cần trả")
            
        try:
            paid_col_index = header.index("Khách đã trả")
        except ValueError:
            missing_columns.append("Khách đã trả")
        
        if missing_columns:
            raise ValueError(f"File danhsachhoadon thiếu cột cần thiết: {', '.join(missing_columns)}")

        # Tạo workbook mới cho kết quả
        output_workbook = Workbook()
        output_sheet = output_workbook.active
        output_sheet.append(["STT", "Tên Khách", "Tổng Tiền", "Tiền mặt", "Chuyển Khoản", "Ship Tuấn", "Ship"])

        # Định dạng header
        for cell in output_sheet[1]:
            apply_cell_style(cell, font=bold_font, alignment=center_alignment, border=thin_border)

        # Xử lý và thêm dữ liệu
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), 2):
            customer = row[customer_col_index].value
            total = row[total_col_index].value
            paid = row[paid_col_index].value

            # Kiểm tra kiểu dữ liệu
            if not isinstance(total, (int, float)) or not isinstance(paid, (int, float)):
                raise ValueError(f"Dữ liệu không hợp lệ ở dòng {row_idx} (cột 'Khách cần trả' và 'Khách đã trả' phải là số).")

            cash = paid if paid > 0 else 0
            transfer = total - cash if cash == 0 else 0

            # Thêm hàng mới vào sheet
            output_sheet.append([row_idx - 1, customer, total, cash, transfer, None, None])
            
            # Căn giữa và định dạng
            for col_idx, cell in enumerate(output_sheet[row_idx], 1):
                apply_cell_style(cell, font=font_style, border=thin_border)
                if col_idx != 2:  # Bỏ qua cột Tên Khách
                    cell.alignment = Alignment(horizontal='center')

        # Thêm dòng tổng
        total_row = output_sheet.max_row + 1
        output_sheet.cell(row=total_row, column=1, value="Tổng")
        
        # Tính tổng cho các cột
        for col_idx in range(3, 8):
            col_letter = get_column_letter(col_idx)
            output_sheet.cell(row=total_row, column=col_idx, value=f"=SUM({col_letter}2:{col_letter}{total_row-1})")
            
        # Định dạng dòng tổng
        for col_idx in range(1, output_sheet.max_column + 1):
            cell = output_sheet.cell(row=total_row, column=col_idx)
            apply_cell_style(cell, font=bold_font, border=thin_border)
            if col_idx > 1:  # Căn giữa tất cả trừ ô đầu tiên (STT)
                cell.alignment = Alignment(horizontal='center')

        # Định dạng số cho các cột tiền tệ
        for col_letter in ["C", "D", "E", "F", "G"]:
            for row in range(2, output_sheet.max_row + 1):
                apply_cell_style(output_sheet[f"{col_letter}{row}"], number_format="#,##0")

        # Điều chỉnh các thuộc tính format
        auto_adjust_column_width(output_sheet)
        
        # Đặt chiều cao hàng
        for row in output_sheet.iter_rows():
            output_sheet.row_dimensions[row[0].row].height = 30

        # Thêm filter
        output_sheet.auto_filter.ref = output_sheet.dimensions

        # Lưu file
        output_workbook.save(output_file_path)
        return output_file_path

    except Exception as e:
        logger.error(f"Lỗi khi xử lý file Excel: {e}")
        return None

def process_multiple_invoice_files(input_file_paths, output_file_path):
    """Xử lý nhiều file hóa đơn và tạo báo cáo tổng hợp."""
    try:
        # Mở file Excel mẫu từ base64
        excel_template_binary = base64.b64decode(EXCEL_TEMPLATE_BASE64)
        output_workbook = load_workbook(BytesIO(excel_template_binary))
        output_sheet = output_workbook.active

        # Điền ngày, tháng, năm hiện tại
        now = datetime.now()
        output_sheet.cell(row=1, column=5, value=now.day)      # Ô E1 (ngày)
        output_sheet.cell(row=1, column=7, value=now.month)    # Ô G1 (tháng)
        output_sheet.cell(row=1, column=9, value=now.year)     # Ô I1 (năm)

        row_num = 11  # Bắt đầu ghi từ dòng thứ 11
        totals = {
            'khach_can_tra': 0,
            'khach_da_tra': 0,
            'gia_tri': 0
        }

        # Lưu thông tin các cột thiếu từ tất cả files
        missing_columns_info = []

        for file_path in input_file_paths:
            file_missing_info = process_single_file(file_path, output_sheet, row_num, totals)
            if file_missing_info:
                missing_columns_info.extend(file_missing_info)
            # Cập nhật row_num sau khi xử lý file
            if "row_num" in locals():
                row_num = locals()["row_num"]

        # Xóa các dòng trống trong phạm vi từ dòng 11 đến dòng 30
        deleted_count, total_chi_row = remove_empty_rows(output_sheet, 11, 30)
        logger.info(f"Đã xóa {deleted_count} dòng trống")

        # Ghi giá trị tổng hợp
        update_summary_values(output_sheet, totals, total_chi_row)

        # Lưu file
        output_workbook.save(output_file_path)

        # Trả về cả file path và thông tin missing columns
        return {
            'file_path': output_file_path,
            'missing_columns_info': missing_columns_info
        }

    except Exception as e:
        logger.error(f"Lỗi khi xử lý nhiều file: {e}")
        return None

def process_single_file(file_path, output_sheet, row_num, totals):
    """Xử lý một file đơn trong quá trình tổng hợp nhiều file."""
    try:
        workbook = load_workbook(filename=file_path, data_only=True)
        sheet = workbook.active
        header = [cell.value for cell in sheet[1]]
        
        missing_info = []
        
        # Detect file type dựa vào tên file thay vì header để track missing columns
        file_name = os.path.basename(file_path).lower()
        
        if file_name.startswith("danhsachhoadon_"):
            # File hóa đơn - luôn gọi process_hoa_don_file để track missing columns
            missing_info = process_hoa_don_file(sheet, header, totals)
        elif file_name.startswith("soquy_"):
            # File sổ quỹ - luôn gọi process_thu_chi_file để track missing columns
            new_row_num, soquy_missing_info = process_thu_chi_file(sheet, header, output_sheet, row_num, totals)
            # Cập nhật row_num cho lần sử dụng tiếp theo
            locals()["row_num"] = new_row_num
            missing_info = soquy_missing_info
        else:
            # Fallback: detect bằng header như trước đây
            if "Khách hàng" in header and "Khách cần trả" in header and "Khách đã trả" in header:
                missing_info = process_hoa_don_file(sheet, header, totals)
            elif "Mã phiếu" in header and "Loại thu chi" in header and "Giá trị" in header:
                new_row_num, soquy_missing_info = process_thu_chi_file(sheet, header, output_sheet, row_num, totals)
                # Cập nhật row_num cho lần sử dụng tiếp theo
                locals()["row_num"] = new_row_num
                missing_info = soquy_missing_info
            else:
                logger.warning(f"Bỏ qua file {file_path} do không xác định được loại file.")
            
        return missing_info
        
    except Exception as e:
        logger.error(f"Lỗi khi xử lý file {file_path}: {e}")
        return []

def process_hoa_don_file(sheet, header, totals):
    """Xử lý dữ liệu từ file hóa đơn."""
    try:
        # Danh sách lưu các cột thiếu
        missing_columns = []
        
        # Kiểm tra các cột bắt buộc
        try:
            customer_col_index = header.index("Khách hàng")
        except ValueError:
            missing_columns.append("Khách hàng")
            
        try:
            total_col_index = header.index("Khách cần trả")
        except ValueError:
            missing_columns.append("Khách cần trả")
            
        try:
            paid_col_index = header.index("Khách đã trả")
        except ValueError:
            missing_columns.append("Khách đã trả")
        
        # Nếu thiếu cột bắt buộc, không thể xử lý nhưng vẫn trả về missing info
        if missing_columns:
            missing_info = [f"File danhsachhoadon thiếu cột: {', '.join(missing_columns)}"]
            return missing_info
        
        # Xử lý dữ liệu nếu có đủ cột
        for row in sheet.iter_rows(min_row=2):
            total_value = row[total_col_index].value
            paid_value = row[paid_col_index].value
            
            totals['khach_can_tra'] += float(total_value) if total_value is not None else 0
            totals['khach_da_tra'] += float(paid_value) if paid_value is not None else 0
            
        return []  # Không có missing columns
        
    except ValueError as e:
        logger.error(f"Lỗi định dạng trong file hóa đơn: {e}")
        return []

def process_thu_chi_file(sheet, header, output_sheet, row_num, totals):
    """Xử lý dữ liệu từ file thu chi."""
    try:
        # Tìm các cột bắt buộc
        column_indices = {
            'ma_phieu': header.index("Mã phiếu"),
            'loai_thu_chi': header.index("Loại thu chi"),
            'nguoi_nop_nhan': header.index("Người nộp/nhận"),
            'gia_tri': header.index("Giá trị")
        }
        
        # Danh sách lưu các cột thiếu
        missing_columns = []
        
        # Tìm cột "Ghi chú" (optional)
        try:
            column_indices['ghi_chu'] = header.index("Ghi chú")
            logger.info("Đã tìm thấy cột 'Ghi chú' trong file soquy")
        except ValueError:
            column_indices['ghi_chu'] = None
            missing_columns.append("Ghi chú")
            logger.info("Không tìm thấy cột 'Ghi chú' trong file soquy - sẽ bỏ qua cột này")
        
        for row in sheet.iter_rows(min_row=2):
            ma_phieu = row[column_indices['ma_phieu']].value
            if ma_phieu is not None:
                # Ghi dữ liệu vào sheet đầu ra
                output_sheet.cell(row=row_num, column=2, value=ma_phieu)  # Mã phiếu
                output_sheet.cell(row=row_num, column=3, value=row[column_indices['loai_thu_chi']].value)  # Nội dung
                output_sheet.cell(row=row_num, column=5, value=row[column_indices['nguoi_nop_nhan']].value)  # Người nộp
                
                # Ghi cột "Ghi chú" nếu có, nếu không thì để trống
                if column_indices['ghi_chu'] is not None:
                    output_sheet.cell(row=row_num, column=7, value=row[column_indices['ghi_chu']].value)  # Ghi chú
                else:
                    output_sheet.cell(row=row_num, column=7, value="")  # Ghi chú trống
                
                output_sheet.cell(row=row_num, column=9, value=row[column_indices['gia_tri']].value)  # Số tiền
                
                row_num += 1
            
            gia_tri = row[column_indices['gia_tri']].value
            if gia_tri is not None:
                totals['gia_tri'] += abs(float(gia_tri))
        
        # Tạo thông báo về cột thiếu nếu có
        missing_info = []
        if missing_columns:
            missing_info.append(f"File soquy thiếu cột: {', '.join(missing_columns)}")
                
        return row_num, missing_info
    except ValueError as e:
        logger.error(f"Lỗi định dạng trong file thu chi: {e}")
        return row_num, []

def remove_empty_rows(sheet, start_row, end_row):
    """Xóa các dòng trống trong phạm vi từ start_row đến dòng trước 'Tổng chi:'.

    Returns:
        tuple: (số dòng đã xóa, vị trí dòng 'Tổng chi:' sau khi xóa)
    """
    # Tìm dòng "Tổng chi:" trước để không xóa các dòng template
    # Kiểm tra nhiều cột vì sau khi merge/unmerge text có thể nằm ở C, D, hoặc E
    total_chi_row_before = None
    for row_idx in range(start_row, sheet.max_row + 1):
        # Kiểm tra cột C, D, E để tìm "Tổng chi:"
        found = False
        for col_idx in [3, 4, 5]:  # Cột C, D, E
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value and "Tổng chi" in str(cell_value):
                total_chi_row_before = row_idx
                logger.info(f"Tìm thấy dòng 'Tổng chi:' tại dòng {total_chi_row_before} cột {col_idx} (trước khi xóa)")
                found = True
                break
        if found:
            break

    # Nếu tìm thấy "Tổng chi:", chỉ xóa từ start_row đến trước dòng đó
    if total_chi_row_before:
        end_row = total_chi_row_before - 1
        logger.info(f"Sẽ xóa dòng trống từ {start_row} đến {end_row}")

    # Duyệt từ dưới lên trên để không bị lỗi index khi xóa
    rows_to_delete = []

    for row_idx in range(start_row, end_row + 1):
        # Kiểm tra xem dòng có dữ liệu không (kiểm tra các cột B, C, E, G, I)
        has_data = False
        for col_idx in [2, 3, 5, 7, 9]:  # Cột B, C, E, G, I
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value is not None and str(cell_value).strip() != "":
                has_data = True
                break

        if not has_data:
            rows_to_delete.append(row_idx)

    # Xóa các dòng từ dưới lên trên
    deleted_count = 0
    for row_idx in reversed(rows_to_delete):
        sheet.delete_rows(row_idx, 1)
        deleted_count += 1
        logger.info(f"Đã xóa dòng trống: {row_idx}")

    # Tính vị trí mới của dòng "Tổng chi:" sau khi xóa
    total_chi_row_after = None
    if total_chi_row_before:
        total_chi_row_after = total_chi_row_before - deleted_count
        logger.info(f"Vị trí dòng 'Tổng chi:' sau khi xóa: {total_chi_row_after}")

    return deleted_count, total_chi_row_after

def update_summary_values(sheet, totals, total_chi_row=None):
    """Cập nhật các giá trị tổng hợp vào file báo cáo.

    Args:
        sheet: Excel worksheet
        totals: Dictionary chứa các tổng
        total_chi_row: Vị trí dòng 'Tổng chi:' (sau khi xóa dòng trống)
    """
    sheet.cell(row=3, column=3, value=totals['khach_can_tra'])  # Doanh thu
    sheet.cell(row=4, column=3, value=totals['khach_da_tra'])  # Tiền mặt
    sheet.cell(row=5, column=3, value=totals['khach_can_tra'] - totals['khach_da_tra'])  # Chuyển khoản

    if total_chi_row:
        # Cập nhật công thức tổng chi tại cột I
        sheet.cell(row=total_chi_row, column=9, value=f"=SUM(I11:I{total_chi_row-1})*-1")
        logger.info(f"Đã cập nhật công thức I{total_chi_row} = SUM(I11:I{total_chi_row-1})*-1")

        # Cập nhật C7 (Phiếu chi) tham chiếu đến I(dòng Tổng chi)
        sheet.cell(row=7, column=3, value=f"=I{total_chi_row}")
        logger.info(f"Đã cập nhật C7 = I{total_chi_row}")

        # Unmerge các cells cũ trước (nếu có) để tránh conflict
        try:
            # Thử unmerge từng vùng có thể bị merge
            for start_col in range(3, 9):  # C đến H
                for end_col in range(start_col, 9):
                    try:
                        sheet.unmerge_cells(start_row=total_chi_row, start_column=start_col,
                                          end_row=total_chi_row, end_column=end_col)
                    except:
                        pass
        except:
            pass

        # Merge cells cho dòng "Tổng chi:" từ C đến H (CDEFGH)
        sheet.merge_cells(start_row=total_chi_row, start_column=3, end_row=total_chi_row, end_column=8)
        logger.info(f"Đã merge cells C{total_chi_row}:H{total_chi_row} cho 'Tổng chi:'")
    else:
        # Fallback: không tìm thấy "Tổng chi:"
        logger.warning("Không tìm thấy dòng 'Tổng chi:', sử dụng giá trị mặc định")
        sheet.cell(row=7, column=3, value=f"=I31")
        sheet.cell(row=31, column=9, value=f"=SUM(I11:I30)*-1")

    # Tồn quỹ = Tiền mặt - Phiếu chi
    sheet.cell(row=8, column=3, value="=C4-C7")

    # Tìm dòng "Số tiền bàn giao:" để cập nhật tồn quỹ
    # Dòng này thường nằm sau dòng "Tổng chi:" 2 dòng (sau khi xóa có thể là 1-2 dòng)
    search_start = total_chi_row + 1 if total_chi_row else 11

    found_ban_giao = False
    for row_idx in range(search_start, min(search_start + 10, sheet.max_row + 1)):
        # Kiểm tra nhiều cột vì merge cells: A, B, C, D, E
        for col_idx in [1, 2, 3, 4, 5]:  # Cột A, B, C, D, E
            cell_value = sheet.cell(row=row_idx, column=col_idx).value

            # Tìm "Số tiền bàn giao:" (có dấu)
            if cell_value and "Số tiền bàn giao" in str(cell_value):
                # Ghi giá trị =C8 vào cột C
                sheet.cell(row=row_idx, column=3, value=f"=C8")

                # Unmerge các cells cũ trước (nếu có) để tránh conflict
                try:
                    # Thử unmerge từng vùng có thể bị merge
                    for start_col in range(3, 10):  # C đến I
                        for end_col in range(start_col, 10):
                            try:
                                sheet.unmerge_cells(start_row=row_idx, start_column=start_col,
                                                  end_row=row_idx, end_column=end_col)
                            except:
                                pass
                except:
                    pass

                # Merge cells cho ô giá trị từ C đến I (CDEFGHI)
                # Cột B (text "Số tiền bàn giao:") không merge
                sheet.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=9)
                logger.info(f"Đã cập nhật dòng {row_idx} 'Số tiền bàn giao:' = C8 và merge C{row_idx}:I{row_idx}")

                found_ban_giao = True
                break
        if found_ban_giao:
            break

    if not found_ban_giao:
        logger.warning(f"Không tìm thấy dòng 'Số tiền bàn giao:' sau dòng Tổng chi")

def process_product_file(input_file_path):
    """Xử lý file sản phẩm và trả về danh sách sản phẩm theo nhóm."""
    try:
        workbook = load_workbook(filename=input_file_path, data_only=True)
        sheet = workbook.active
        
        result = extract_product_data(sheet)
        return format_product_data(result)
        
    except Exception as e:
        logger.error(f"Lỗi khi xử lý file sản phẩm: {e}")
        return None

def extract_product_data(sheet):
    """Trích xuất dữ liệu sản phẩm từ sheet."""
    # Tìm vị trí các cột
    header = [cell.value for cell in sheet[1]]
    group_col_index = header.index("Nhóm hàng(3 Cấp)")
    product_name_col_index = header.index("Tên hàng")
    stock_col_index = header.index("Tồn kho")
    
    # Lọc dữ liệu - hiển thị tất cả các nhóm, không chỉ nhóm cụ thể
    filtered_data = {}
    
    for row in sheet.iter_rows(min_row=2):
        group = row[group_col_index].value
        product_name = row[product_name_col_index].value
        stock = row[stock_col_index].value
        
        if stock != 0:  # Hiển thị cả sản phẩm có tồn kho âm và dương, bỏ qua chỉ = 0
            if group not in filtered_data:
                filtered_data[group] = []
            filtered_data[group].append(f"- {product_name}: {stock}")
    
    # Sắp xếp sản phẩm theo alphabet tiếng Việt
    for group in filtered_data:
        filtered_data[group] = sorted(filtered_data[group], key=locale.strxfrm)
    
    sorted_groups = sorted(filtered_data.keys(), key=locale.strxfrm)
    
    return {
        'filtered_data': filtered_data,
        'sorted_groups': sorted_groups
    }

def format_product_data(data):
    """Định dạng dữ liệu sản phẩm thành chuỗi kết quả."""
    filtered_data = data['filtered_data']
    sorted_groups = data['sorted_groups']
    
    output_string = "Danh sách sản phẩm có hàng tồn khác 0 (bao gồm cả tồn kho âm) :\n\n"
    for group in sorted_groups:
        if filtered_data[group]:  # Chỉ hiển thị nhóm có sản phẩm
            output_string += f"Nhóm: {group}\n"
            for product in filtered_data[group]:
                output_string += f"{product}\n"
            output_string += "\n"
    
    return output_string

def process_excel_file_updated(file_path):
    """Xử lý file Excel và trả về dữ liệu định dạng có cấu trúc."""
    try:
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active
        
        # Tìm vị trí các cột
        header = [cell.value for cell in sheet[1]]
        
        # Danh sách lưu các cột thiếu
        missing_columns = []
        
        # Kiểm tra các cột bắt buộc
        try:
            group_col_index = header.index("Nhóm hàng(3 Cấp)")
        except ValueError:
            missing_columns.append("Nhóm hàng(3 Cấp)")
            return f"Lỗi: File danhsachsanpham thiếu cột bắt buộc 'Nhóm hàng(3 Cấp)'"
            
        try:
            product_name_col_index = header.index("Tên hàng")
        except ValueError:
            missing_columns.append("Tên hàng")
            return f"Lỗi: File danhsachsanpham thiếu cột bắt buộc 'Tên hàng'"
            
        try:
            stock_col_index = header.index("Tồn kho")
        except ValueError:
            missing_columns.append("Tồn kho")
            return f"Lỗi: File danhsachsanpham thiếu cột bắt buộc 'Tồn kho'"
        
        # Tìm cột "Giá vốn" (optional)
        unit_cost_col_index = None
        try:
            unit_cost_col_index = header.index("Giá vốn")
            logger.info(f"Đã tìm thấy cột Giá vốn tại vị trí {unit_cost_col_index}")
        except ValueError:
            # Thử tìm các biến thể khác của cột giá vốn
            for i, col_name in enumerate(header):
                if col_name and isinstance(col_name, str) and 'giá vốn' in col_name.lower():
                    unit_cost_col_index = i
                    logger.info(f"Đã tìm thấy cột giá vốn (tìm mờ) tại vị trí {unit_cost_col_index}: {col_name}")
                    break
            
            if unit_cost_col_index is None:
                missing_columns.append("Giá vốn")
                logger.warning("Không tìm thấy cột 'Giá vốn' - sẽ bỏ qua tính tổng tiền tồn kho")
        
        # Dữ liệu đầu ra
        all_products = []
        
        # Danh sách các nhóm bị loại trừ
        excluded_groups = ["Nước rửa chén"]
        
        # Tìm tất cả các nhóm hàng trong file, ngoại trừ các nhóm bị loại trừ
        all_groups = set()
        for row in sheet.iter_rows(min_row=2):
            group = row[group_col_index].value
            if group and group not in excluded_groups:
                all_groups.add(group)
        
        # Tạo dict lưu trữ sản phẩm theo nhóm
        filtered_data = {group: [] for group in all_groups}
        
        # Dictionary để lưu thông tin giá vốn × tồn kho cho từng sản phẩm
        product_cost_info = {}
        
        # Xử lý dữ liệu
        for row in sheet.iter_rows(min_row=2):
            group = row[group_col_index].value
            product_name = row[product_name_col_index].value
            stock = row[stock_col_index].value
            
            if stock != 0:  # Hiển thị cả sản phẩm có tồn kho âm và dương, bỏ qua chỉ = 0
                all_products.append(f"- {product_name}: {stock}")
                
                # Tính tổng tiền tồn kho = Giá vốn × Tồn kho
                total_cost = 0
                if unit_cost_col_index is not None:
                    unit_cost_value = row[unit_cost_col_index].value
                    if unit_cost_value is not None:
                        try:
                            unit_cost = float(unit_cost_value)
                            total_cost = unit_cost * float(stock)
                        except (ValueError, TypeError):
                            logger.warning(f"Giá vốn hoặc tồn kho không hợp lệ cho sản phẩm '{product_name}': giá vốn={unit_cost_value}, tồn kho={stock}")
                            total_cost = 0
                
                # Lưu thông tin cost cho sản phẩm này
                product_cost_info[product_name] = {
                    "stock": float(stock),
                    "total_cost": total_cost
                }
                
                if group and group not in excluded_groups:
                    filtered_data[group].append(f"- {product_name}: {stock}")
        
        # Sắp xếp dữ liệu
        all_products.sort(key=locale.strxfrm)
        for group in filtered_data:
            filtered_data[group].sort(key=locale.strxfrm)
        
        sorted_groups = sorted(filtered_data.keys(), key=locale.strxfrm)
        
        # Tạo thông báo về cột thiếu nếu có
        missing_info = []
        if missing_columns:
            missing_info.append(f"File danhsachsanpham thiếu cột: {', '.join(missing_columns)}")
        
        return {
            "all_products": all_products,
            "grouped_products": filtered_data,
            "sorted_groups": sorted_groups,
            "product_cost_info": product_cost_info,
            "missing_columns_info": missing_info
        }
        
    except Exception as e:
        logger.error(f"Lỗi khi xử lý file Excel cập nhật: {e}")
        return f"Lỗi khi xử lý file Excel: {e}"

def process_invoice_file(input_file_path, output_file_path):
    """Xử lý file hóa đơn đơn với tracking missing columns."""
    try:
        result_path = process_excel_file(input_file_path, output_file_path)
        if result_path:
            # Thành công - không có missing columns
            return {
                'file_path': result_path,
                'missing_columns_info': []
            }
        else:
            return None
    except ValueError as e:
        # Check nếu là lỗi missing columns
        error_msg = str(e)
        if "thiếu cột cần thiết" in error_msg:
            # Extract missing columns info từ error message
            return {
                'file_path': None,
                'missing_columns_info': [error_msg]
            }
        else:
            return None
    except Exception as e:
        logger.error(f"Lỗi khi xử lý file hóa đơn: {e}")
        return None

def process_purchase_order_detail_file(file_path):
    """Xử lý file Excel chi tiết đơn mua hàng từ KiotViet."""
    try:
        workbook = load_workbook(filename=file_path, data_only=True)
        sheet = workbook.active
        
        # Tìm các cột quan trọng
        header = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
        logger.info(f"Các cột tìm thấy trong file: {header}")
        
        try:
            # Tìm chính xác cột "Tên nhà cung cấp", phân biệt hoa thường
            supplier_col_index = None
            product_name_col_index = None
            quantity_col_index = None
            unit_price_col_index = None
            
            for i, col in enumerate(header):
                if col == "Tên nhà cung cấp":
                    supplier_col_index = i
                    logger.info(f"Đã tìm thấy cột Tên nhà cung cấp chính xác tại vị trí {i}: {col}")
                elif col == "Tên hàng":
                    product_name_col_index = i
                    logger.info(f"Đã tìm thấy cột Tên hàng chính xác tại vị trí {i}: {col}")
                elif col == "Số lượng":
                    quantity_col_index = i
                    logger.info(f"Đã tìm thấy cột Số lượng chính xác tại vị trí {i}: {col}")
                elif col == "Giá nhập":
                    unit_price_col_index = i
                    logger.info(f"Đã tìm thấy cột Giá nhập chính xác tại vị trí {i}: {col}")
            
            # Nếu không tìm thấy, thử tìm cách khác không phân biệt hoa thường
            if supplier_col_index is None:
                supplier_col_index = next((i for i, col in enumerate(header) if col.lower() == "tên nhà cung cấp"), None)
                if supplier_col_index is not None:
                    logger.info(f"Đã tìm thấy cột tên nhà cung cấp (không phân biệt hoa thường) tại vị trí {supplier_col_index}: {header[supplier_col_index]}")
            
            if product_name_col_index is None:
                product_name_col_index = next((i for i, col in enumerate(header) if col.lower() == "tên hàng"), None)
                if product_name_col_index is not None:
                    logger.info(f"Đã tìm thấy cột tên hàng (không phân biệt hoa thường) tại vị trí {product_name_col_index}: {header[product_name_col_index]}")
            
            if quantity_col_index is None:
                quantity_col_index = next((i for i, col in enumerate(header) if col.lower() == "số lượng"), None)
                if quantity_col_index is not None:
                    logger.info(f"Đã tìm thấy cột số lượng (không phân biệt hoa thường) tại vị trí {quantity_col_index}: {header[quantity_col_index]}")
            
            if unit_price_col_index is None:
                unit_price_col_index = next((i for i, col in enumerate(header) if col.lower() == "giá nhập"), None)
                if unit_price_col_index is not None:
                    logger.info(f"Đã tìm thấy cột giá nhập (không phân biệt hoa thường) tại vị trí {unit_price_col_index}: {header[unit_price_col_index]}")
            
            # Nếu vẫn không tìm thấy, thử tìm kiếm mờ
            if supplier_col_index is None:
                supplier_col_index = next((i for i, col in enumerate(header) if "tên nhà cung cấp" in col.lower()), None)
                if supplier_col_index is not None:
                    logger.info(f"Đã tìm thấy cột tên nhà cung cấp (tìm mờ) tại vị trí {supplier_col_index}: {header[supplier_col_index]}")
            
            if product_name_col_index is None:
                product_name_col_index = next((i for i, col in enumerate(header) if "tên hàng" in col.lower()), None)
                if product_name_col_index is not None:
                    logger.info(f"Đã tìm thấy cột tên hàng (tìm mờ) tại vị trí {product_name_col_index}: {header[product_name_col_index]}")
            
            if quantity_col_index is None:
                quantity_col_index = next((i for i, col in enumerate(header) if "số lượng" in col.lower()), None)
                if quantity_col_index is not None:
                    logger.info(f"Đã tìm thấy cột số lượng (tìm mờ) tại vị trí {quantity_col_index}: {header[quantity_col_index]}")
            
            if unit_price_col_index is None:
                unit_price_col_index = next((i for i, col in enumerate(header) if "giá nhập" in col.lower()), None)
                if unit_price_col_index is not None:
                    logger.info(f"Đã tìm thấy cột giá nhập (tìm mờ) tại vị trí {unit_price_col_index}: {header[unit_price_col_index]}")
            
            if supplier_col_index is None or product_name_col_index is None or quantity_col_index is None:
                logger.error("Không tìm thấy một hoặc nhiều cột cần thiết trong file đơn mua hàng")
                logger.error(f"Supplier col: {supplier_col_index}, Product name col: {product_name_col_index}, Quantity col: {quantity_col_index}, Unit price col: {unit_price_col_index}")
                return f"Lỗi: Không tìm thấy các cột cần thiết trong file. Cần có 'Tên nhà cung cấp', 'Tên hàng', 'Số lượng'."
            
            # Chú ý: cột "Giá nhập" là optional, nếu không có thì sẽ skip tính tổng tiền
            if unit_price_col_index is None:
                logger.warning("Không tìm thấy cột 'Giá nhập' - sẽ bỏ qua tính tổng tiền")
        
        except Exception as e:
            logger.error(f"Lỗi khi tìm vị trí các cột: {e}")
            return f"Lỗi khi tìm vị trí các cột: {e}"
        
        # Dictionary lưu trữ dữ liệu theo nhà cung cấp
        suppliers_data = {}
        
        # Duyệt qua các dòng từ dòng thứ 2 (dữ liệu)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            supplier = row[supplier_col_index].value
            product_name = row[product_name_col_index].value
            quantity = row[quantity_col_index].value
            
            # Bỏ qua dòng nếu thiếu thông tin
            if not supplier or not product_name or quantity is None:
                continue
            
            # Chuyển đổi số lượng sang số
            try:
                quantity_num = float(quantity)
                if quantity_num <= 0:
                    continue
            except (ValueError, TypeError):
                logger.warning(f"Số lượng không hợp lệ ở dòng {row_idx}: {quantity}")
                continue
            
            # Lấy giá nhập và tính tổng tiền = giá nhập × số lượng
            total_price = 0
            if unit_price_col_index is not None:
                unit_price_value = row[unit_price_col_index].value
                if unit_price_value is not None:
                    try:
                        unit_price = float(unit_price_value)
                        total_price = unit_price * quantity_num
                    except (ValueError, TypeError):
                        logger.warning(f"Giá nhập không hợp lệ ở dòng {row_idx}: {unit_price_value}")
                        total_price = 0
            
            # Khởi tạo dictionary cho nhà cung cấp nếu chưa có
            if supplier not in suppliers_data:
                suppliers_data[supplier] = {}
            
            # Cộng dồn số lượng và tổng tiền cho sản phẩm
            if product_name in suppliers_data[supplier]:
                suppliers_data[supplier][product_name]["quantity"] += quantity_num
                suppliers_data[supplier][product_name]["total_price"] += total_price
            else:
                suppliers_data[supplier][product_name] = {
                    "quantity": quantity_num,
                    "total_price": total_price
                }
        
        # Sắp xếp kết quả theo tên nhà cung cấp (theo bảng chữ cái tiếng Việt)
        sorted_suppliers = sorted(suppliers_data.keys(), key=locale.strxfrm)
        sorted_result = {supplier: suppliers_data[supplier] for supplier in sorted_suppliers}
        
        # Với mỗi nhà cung cấp, sắp xếp sản phẩm theo tên
        for supplier in sorted_result:
            sorted_products = {k: v for k, v in sorted(sorted_result[supplier].items(), key=lambda item: locale.strxfrm(str(item[0])))}
            sorted_result[supplier] = sorted_products
        
        return sorted_result
    
    except Exception as e:
        logger.error(f"Lỗi khi xử lý file đơn mua hàng: {e}")
        return f"Lỗi khi xử lý file đơn mua hàng: {e}"

# ============================================================================
# BOT HANDLERS (từ bot.py)
# ============================================================================

# Decorator kiểm tra quyền truy cập
def restricted(func):
    """Decorator để hạn chế truy cập bot."""
    @wraps(func)
    async def wrapped(update: Update, context: ContextTypes.DEFAULT_TYPE, *args, **kwargs):
        user_id = update.effective_user.id
        
        # Nếu ALLOWED_USERS trống, cho phép tất cả
        if not ALLOWED_USERS:
            return await func(update, context, *args, **kwargs)
        
        # Kiểm tra user_id có trong danh sách cho phép
        if user_id not in ALLOWED_USERS:
            logger.warning(f"Từ chối truy cập từ user {user_id}")
            await update.message.reply_text(
                "❌ Bạn không có quyền sử dụng bot này.\n"
                f"User ID của bạn: {user_id}"
            )
            return
        
        return await func(update, context, *args, **kwargs)
    
    return wrapped

# Command handlers
@restricted
async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handler cho lệnh /start."""
    user = update.effective_user
    welcome_message = (
        f"👋 Xin chào {user.first_name}!\n\n"
        "🤖 Bot Xử Lý Excel Tự Động\n\n"
        "📋 Các loại file được hỗ trợ:\n\n"
        "• Danh sách hóa đơn (danhsachhoadon_*.xlsx)\n"
        "  → Tạo báo cáo tổng hợp hóa đơn\n\n"
        "• Sổ quỹ (soquy_*.xlsx)\n"
        "  → Gộp với hóa đơn thành báo cáo thu chi\n\n"
        "• Danh sách sản phẩm (danhsachsanpham_*.xlsx)\n"
        "  → Lọc sản phẩm tồn kho ≠ 0\n\n"
        "• Chi tiết đơn đặt hàng (danhsachchitietdathang_*.xlsx)\n"
        "  → Nhóm theo nhà cung cấp\n\n"
        "💡 Cách sử dụng:\n"
        "1. Gửi file Excel vào chat\n"
        "2. Bot sẽ tự động xử lý\n"
        "3. Nhận kết quả ngay lập tức!\n\n"
        "📌 Lưu ý: Tên file phải đúng định dạng để bot nhận diện."
    )
    
    await update.message.reply_text(welcome_message)

@restricted
async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handler cho lệnh /help."""
    help_text = (
        "📚 Hướng Dẫn Sử Dụng Bot\n\n"
        "1️⃣ File Danh Sách Hóa Đơn:\n"
        "• Tên file: danhsachhoadon_*.xlsx\n"
        "• Cần có cột: Khách hàng, Khách cần trả, Khách đã trả\n"
        "• Kết quả: File Excel với tổng tiền, tiền mặt, chuyển khoản\n\n"
        "2️⃣ File Sổ Quỹ:\n"
        "• Tên file: soquy_*.xlsx\n"
        "• Cần có cột: Mã phiếu, Loại thu chi, Người nộp/nhận, Giá trị\n"
        "• Kết quả: Gộp với file hóa đơn thành báo cáo tổng hợp\n\n"
        "3️⃣ File Danh Sách Sản Phẩm:\n"
        "• Tên file: danhsachsanpham_*.xlsx\n"
        "• Cần có cột: Nhóm hàng(3 Cấp), Tên hàng, Tồn kho\n"
        "• Kết quả: Danh sách sản phẩm nhóm theo danh mục\n\n"
        "4️⃣ File Chi Tiết Đơn Đặt Hàng:\n"
        "• Tên file: danhsachchitietdathang_*.xlsx\n"
        "• Cần có cột: Tên nhà cung cấp, Tên hàng, Số lượng\n"
        "• Kết quả: Danh sách nhóm theo nhà cung cấp\n\n"
        "🔄 Gộp File:\n"
        "Gửi 1 file danhsachhoadon + 1 file soquy → Bot tự động tổng hợp!\n\n"
        "📞 Lệnh hỗ trợ:\n"
        "/start - Khởi động bot\n"
        "/help - Xem hướng dẫn\n"
        "/clear - Xóa dữ liệu tạm\n"
        "/tinhluong - Gửi file bảng lương"
    )
    
    await update.message.reply_text(help_text)

@restricted
async def clear_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Xóa dữ liệu tạm trong context."""
    # Cleanup temp directories nếu có
    for key in ['invoice_tempdir', 'soquy_tempdir', 'combine_tempdir']:
        tempdir = context.user_data.get(key)
        if tempdir and os.path.exists(tempdir):
            try:
                shutil.rmtree(tempdir)
                logger.info(f"Cleaned up {key}: {tempdir}")
            except Exception as e:
                logger.error(f"Error cleaning {key}: {e}")
    
    # Clear user data
    context.user_data.clear()
    
    await update.message.reply_text("✅ Đã xóa tất cả dữ liệu tạm!")

@restricted
async def tinhluong_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Gửi file bảng lương từ biến môi trường BANGLUONG."""
    await update.message.reply_text("⏳ Đang chuẩn bị file bảng lương...")
    logger.info(f"User {update.effective_user.id} yêu cầu file bảng lương.")
    
    temp_payroll_dir = None
    try:
        # Lấy dữ liệu bảng lương từ biến môi trường
        if not BANGLUONG:
            logger.warning("BANGLUONG env var not found.")
            await update.message.reply_text("❌ Lỗi: Không tìm thấy dữ liệu bảng lương (BANGLUONG).")
            return

        # Giải mã dữ liệu base64
        try:
            excel_binary = base64.b64decode(BANGLUONG)
        except (base64.binascii.Error, TypeError) as decode_error:
            logger.error(f"Failed to decode BANGLUONG base64: {decode_error}")
            await update.message.reply_text("❌ Lỗi: Dữ liệu bảng lương bị lỗi.")
            return
            
        # Tạo file tạm và lưu dữ liệu
        temp_payroll_dir = tempfile.mkdtemp(prefix="payroll_")
        file_name = f"BangLuong_{datetime.now().strftime('%d%m')}.xlsx"
        file_path = os.path.join(temp_payroll_dir, file_name)
        
        with open(file_path, 'wb') as f:
            f.write(excel_binary)
        
        logger.info(f"Created temp payroll file: {file_path}")
        
        # Gửi file đến người dùng
        with open(file_path, 'rb') as f:
            await update.message.reply_document(
                document=f,
                filename=file_name,
                caption="💰 Bảng lương đã sẵn sàng!"
            )
        
        logger.info(f"Sent payroll file '{file_name}' to user {update.effective_user.id}")

    except Exception as e:
        logger.error(f"Error in /tinhluong: {e}", exc_info=True)
        await update.message.reply_text(f"❌ Lỗi không mong muốn khi xử lý bảng lương: {str(e)[:100]}")
        
    finally:
        # Dọn dẹp thư mục tạm
        if temp_payroll_dir and os.path.exists(temp_payroll_dir):
            try:
                shutil.rmtree(temp_payroll_dir)
                logger.info(f"Cleaned payroll temp: {temp_payroll_dir}")
            except Exception as ce:
                logger.error(f"Error cleaning payroll temp {temp_payroll_dir}: {ce}")

# File handlers
@restricted
async def handle_excel_file(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Xử lý file Excel được gửi vào."""
    if not update.message or not update.message.document:
        logger.warning("handle_excel_file được gọi nhưng không có document.")
        return

    file = await update.message.document.get_file()
    file_name = update.message.document.file_name
    
    # Kiểm tra kích thước file
    file_size = update.message.document.file_size
    if file_size:
        file_size_mb = file_size / (1024 * 1024)
        if file_size_mb > MAX_FILE_SIZE_MB:
            await update.message.reply_text(
                f"❌ File '{file_name}' quá lớn ({file_size_mb:.1f}MB). "
                f"Giới hạn: {MAX_FILE_SIZE_MB}MB. Vui lòng nén hoặc chia nhỏ file."
            )
            return
    
    temp_dir = tempfile.mkdtemp(prefix="telegram_dl_")
    file_path = os.path.join(temp_dir, file_name)
    should_cleanup_immediately = False

    try:
        await file.download_to_drive(file_path)
        logger.info(f"Downloaded file '{file_name}' to '{file_path}'")

        file_name_lower = file_name.lower()

        # Phát hiện loại file và xử lý
        if file_name_lower.startswith("danhsachhoadon_"):
            await handle_danhsachhoadon_file(update, context, file_path, file_name, temp_dir)
            
        elif file_name_lower.startswith("soquy_"):
            await handle_soquy_file(update, context, file_path, file_name, temp_dir)
            
        elif file_name_lower.startswith("danhsachsanpham_"):
            await handle_danhsachsanpham_file(update, context, file_path, file_name)
            should_cleanup_immediately = True
            
        elif file_name_lower.startswith("danhsachchitietdathang_"):
            await handle_danhsachchitietdathang_file(update, context, file_path, file_name)
            should_cleanup_immediately = True
        
        else:
            await update.message.reply_text(
                f"❌ File '{file_name}' không được nhận diện.\n\n"
                "Vui lòng đặt tên file theo định dạng:\n"
                "• danhsachhoadon_*.xlsx\n"
                "• soquy_*.xlsx\n"
                "• danhsachsanpham_*.xlsx\n"
                "• danhsachchitietdathang_*.xlsx"
            )
            should_cleanup_immediately = True

    except Exception as e:
        logger.error(f"Lỗi khi xử lý file {file_name}: {e}", exc_info=True)
        await update.message.reply_text(
            f"❌ Đã xảy ra lỗi khi xử lý file '{file_name}'.\n"
            f"Chi tiết: {str(e)[:100]}..."
        )
        should_cleanup_immediately = True
        
    finally:
        if should_cleanup_immediately and os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir)
                logger.info(f"Cleaned up temp directory: {temp_dir}")
            except Exception as cleanup_error:
                logger.error(f"Error cleaning up {temp_dir}: {cleanup_error}")

async def handle_danhsachhoadon_file(update, context, file_path, file_name, temp_dir):
    """Xử lý file danh sách hóa đơn."""
    status_msg = await update.message.reply_text("⏳ Đang xử lý file danh sách hóa đơn...")
    
    try:
        # Kiểm tra có file soquy đang chờ không
        has_soquy = context.user_data.get('soquy_file') is not None
        
        if has_soquy:
            # Nếu có file soquy → Lưu file và tổng hợp
            context.user_data['invoice_file'] = file_path
            context.user_data['invoice_tempdir'] = temp_dir
            
            await status_msg.edit_text("✅ Đã nhận file hóa đơn!")
            await auto_combine_reports(update, context)
        else:
            # Nếu KHÔNG có file soquy → Xử lý riêng lẻ, KHÔNG lưu vào context
            output_path = os.path.join(temp_dir, f"processed_{file_name}")
            result = process_invoice_file(file_path, output_path)
            
            if result and result.get('file_path'):
                # Gửi file kết quả riêng lẻ
                with open(result['file_path'], 'rb') as f:
                    await update.message.reply_document(
                        document=f,
                        filename=f"KetQua_{file_name}",
                        caption=f"✅ Đã xử lý file: {file_name}"
                    )
                
                await status_msg.edit_text("✅ Xử lý file danh sách hóa đơn thành công!")
                
                # KHÔNG lưu vào context vì đã xử lý xong riêng lẻ
                # Cleanup temp dir ngay
                if os.path.exists(temp_dir):
                    try:
                        shutil.rmtree(temp_dir)
                        logger.info(f"Cleaned up temp dir after standalone processing: {temp_dir}")
                    except Exception as cleanup_error:
                        logger.error(f"Error cleaning temp dir: {cleanup_error}")
            else:
                # Xử lý lỗi
                missing_info = result.get('missing_columns_info', []) if result else []
                error_msg = "❌ Không thể xử lý file.\n"
                if missing_info:
                    error_msg += f"Lỗi: {', '.join(missing_info)}"
                await status_msg.edit_text(error_msg)
            
    except Exception as e:
        logger.error(f"Lỗi xử lý file hóa đơn: {e}", exc_info=True)
        await status_msg.edit_text(f"❌ Lỗi: {str(e)[:100]}")

async def handle_soquy_file(update, context, file_path, file_name, temp_dir):
    """Xử lý file sổ quỹ."""
    status_msg = await update.message.reply_text("⏳ Đang lưu file sổ quỹ...")
    
    try:
        # Lưu file vào context (chỉ lưu 1 file)
        context.user_data['soquy_file'] = file_path
        context.user_data['soquy_tempdir'] = temp_dir
        
        await status_msg.edit_text("✅ Đã lưu file sổ quỹ!")
        
        # Thông báo chờ file hóa đơn MỚI
        # KHÔNG tổng hợp với file invoice cũ (nếu có)
        await update.message.reply_text(
            "💡 Đã lưu file sổ quỹ.\n"
            "Hãy gửi file danhsachhoadon_*.xlsx để tạo báo cáo tổng hợp!"
        )
            
    except Exception as e:
        logger.error(f"Lỗi xử lý file sổ quỹ: {e}", exc_info=True)
        await status_msg.edit_text(f"❌ Lỗi: {str(e)[:100]}")

async def handle_danhsachsanpham_file(update, context, file_path, file_name):
    """Xử lý file danh sách sản phẩm."""
    status_msg = await update.message.reply_text("⏳ Đang xử lý file danh sách sản phẩm...")
    
    try:
        result_data = process_excel_file_updated(file_path)
        
        if isinstance(result_data, dict):
            # Tạo message từ grouped_products
            output_string = "📦 E gửi danh Sách Sản Phẩm Tồn Kho ≠ 0\n\n"
            
            for group in result_data.get('sorted_groups', []):
                products = result_data['grouped_products'].get(group, [])
                if products:
                    output_string += f"Nhóm: {group}\n"
                    for product in products:
                        output_string += f"{product}\n"
                    output_string += "\n"
            
            # Kiểm tra missing columns
            missing_info = result_data.get('missing_columns_info', [])
            if missing_info:
                output_string += f"\n⚠️ Cảnh báo:\n{', '.join(missing_info)}\n"
            
            # Gửi kết quả (chia nhỏ nếu quá dài)
            if len(output_string) > 4000:
                # Chia thành nhiều message
                parts = [output_string[i:i+4000] for i in range(0, len(output_string), 4000)]
                for part in parts:
                    await update.message.reply_text(part)
            else:
                await update.message.reply_text(output_string)
            
            await status_msg.edit_text("✅ Xử lý file danh sách sản phẩm thành công!")
        else:
            await status_msg.edit_text(f"❌ Lỗi: {result_data}")
            
    except Exception as e:
        logger.error(f"Lỗi xử lý file sản phẩm: {e}", exc_info=True)
        await status_msg.edit_text(f"❌ Lỗi: {str(e)[:100]}")

async def handle_danhsachchitietdathang_file(update, context, file_path, file_name):
    """Xử lý file chi tiết đơn đặt hàng."""
    status_msg = await update.message.reply_text("⏳ Đang xử lý file chi tiết đơn đặt hàng...")
    
    try:
        result_data = process_purchase_order_detail_file(file_path)
        
        if isinstance(result_data, dict):
            # Tạo message từ suppliers_data
            output_string = "🛒 Chi Tiết Đơn Đặt Hàng Theo Nhà Cung Cấp\n\n"
            
            for supplier, products in result_data.items():
                output_string += f"{supplier}:\n"
                total_supplier_amount = 0
                
                for product_name, info in products.items():
                    quantity = info.get('quantity', 0)
                    total_price = info.get('total_price', 0)
                    total_supplier_amount += total_price
                    
                    if total_price > 0:
                        output_string += f"• {product_name}: {quantity} (Tổng: {total_price:,.0f}đ)\n"
                    else:
                        output_string += f"• {product_name}: {quantity}\n"
                
                if total_supplier_amount > 0:
                    output_string += f"Tổng: {total_supplier_amount:,.0f}đ\n\n"
                else:
                    output_string += "\n"
            
            # Gửi kết quả (chia nhỏ nếu quá dài)
            if len(output_string) > 4000:
                parts = [output_string[i:i+4000] for i in range(0, len(output_string), 4000)]
                for part in parts:
                    await update.message.reply_text(part)
            else:
                await update.message.reply_text(output_string)
            
            await status_msg.edit_text("✅ Xử lý file chi tiết đơn đặt hàng thành công!")
        else:
            await status_msg.edit_text(f"❌ Lỗi: {result_data}")
            
    except Exception as e:
        logger.error(f"Lỗi xử lý file đơn đặt hàng: {e}", exc_info=True)
        await status_msg.edit_text(f"❌ Lỗi: {str(e)[:100]}")

async def auto_combine_reports(update, context):
    """Tự động tổng hợp 1 file hóa đơn + 1 file sổ quỹ."""
    status_msg = await update.message.reply_text("⏳ Đang tổng hợp báo cáo...")
    
    try:
        invoice_file = context.user_data.get('invoice_file')
        soquy_file = context.user_data.get('soquy_file')
        
        if not invoice_file or not soquy_file:
            await status_msg.edit_text("❌ Thiếu file hóa đơn hoặc sổ quỹ!")
            return
        
        # Kiểm tra file tồn tại
        if not os.path.exists(invoice_file) or not os.path.exists(soquy_file):
            await status_msg.edit_text("❌ File không tồn tại!")
            return
        
        logger.info(f"Tự động tổng hợp: {os.path.basename(invoice_file)} + {os.path.basename(soquy_file)}")
        
        # Tạo temp dir cho output
        combine_temp_dir = tempfile.mkdtemp(prefix="combine_")
        context.user_data['combine_tempdir'] = combine_temp_dir
        
        output_file_path = os.path.join(
            combine_temp_dir,
            f"TongHop_{datetime.now().strftime('%d%m%Y_%H%M%S')}.xlsx"
        )
        
        # Gộp 2 files
        all_files = [invoice_file, soquy_file]
        
        # Xử lý
        result = process_multiple_invoice_files(all_files, output_file_path)
        
        if result and result.get('file_path'):
            # Gửi file kết quả
            with open(result['file_path'], 'rb') as f:
                await update.message.reply_document(
                    document=f,
                    filename=os.path.basename(result['file_path']),
                    caption="✅ Báo cáo tổng hợp đã sẵn sàng!"
                )
            
            # Hiển thị warning nếu có missing columns
            missing_info = result.get('missing_columns_info', [])
            if missing_info:
                warning_msg = "⚠️ Cảnh báo:\n" + "\n".join(missing_info)
                await update.message.reply_text(warning_msg)
            
            await status_msg.edit_text("✅ Tổng hợp thành công!")
            
            # Cleanup
            shutil.rmtree(combine_temp_dir)
            context.user_data.clear()
            
            logger.info(f"Đã gửi file tổng hợp: {os.path.basename(result['file_path'])}")
        else:
            await status_msg.edit_text("❌ Không thể tổng hợp báo cáo!")
            
    except Exception as e:
        logger.error(f"Lỗi tổng hợp báo cáo: {e}", exc_info=True)
        await status_msg.edit_text(f"❌ Lỗi: {str(e)[:100]}")

def bot_main():
    """Khởi động bot."""
    if not TELEGRAM_TOKEN:
        logger.error("❌ TELEGRAM_TOKEN không được tìm thấy! Vui lòng kiểm tra file .env")
        return
    
    # Tạo application
    application = Application.builder().token(TELEGRAM_TOKEN).build()
    
    # Đăng ký handlers
    application.add_handler(CommandHandler("start", start_command))
    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(CommandHandler("clear", clear_command))
    application.add_handler(CommandHandler("tinhluong", tinhluong_command))
    
    # Handler cho file Excel
    application.add_handler(MessageHandler(
        filters.Document.FileExtension("xlsx") | filters.Document.FileExtension("xls"),
        handle_excel_file
    ))
    
    # Khởi động bot
    logger.info("🤖 Bot đang khởi động...")
    application.run_polling(allowed_updates=Update.ALL_TYPES)

# ============================================================================
# MAIN ENTRY POINT (từ main.py)
# ============================================================================

def main():
    """Khởi động bot."""
    try:
        logger.info("=" * 50)
        logger.info("🤖 EXCEL BOT - TELEGRAM BOT XỬ LÝ EXCEL")
        logger.info("=" * 50)
        logger.info("📝 Phiên bản: 1.0.0")
        logger.info("📅 Ngày: 02/10/2025")
        logger.info("=" * 50)
        
        # Chạy bot
        bot_main()
        
    except KeyboardInterrupt:
        logger.info("\n⏹️  Bot đã dừng (Ctrl+C)")
        sys.exit(0)
    except Exception as e:
        logger.error(f"❌ Lỗi khi khởi động bot: {e}", exc_info=True)
        sys.exit(1)

if __name__ == "__main__":
    # Thiết lập logging
    logging.basicConfig(
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        level=logging.INFO
    )
    
    main()

